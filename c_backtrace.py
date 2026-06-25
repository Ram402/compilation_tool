#!/usr/bin/env python3
"""
Automatic C function backtracer for AUTOSAR / PDC projects.

Builds a reverse call graph from a source tree and writes caller chains
to an Excel workbook matching the Back_tracing spreadsheet layout.

Usage:
    python c_backtrace.py
        (prompts for source directory and .c file name)

    python c_backtrace.py --src-dir <path> --file PDC_FrunkLamp.c
    python c_backtrace.py --src-dir <path> --file Aswc_Frunk_Lp.c -o Frunk_Lamp_backtrace.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


C_KEYWORDS = {
    "if", "else", "for", "while", "do", "switch", "case", "default", "return",
    "break", "continue", "goto", "sizeof", "typeof", "alignof", "static_assert",
    "struct", "union", "enum", "typedef", "const", "volatile", "register",
    "auto", "extern", "static", "inline", "restrict", "_Bool", "_Complex",
    "_Imaginary", "void", "char", "short", "int", "long", "float", "double",
    "signed", "unsigned", "bool", "true", "false", "NULL", "stdin", "stdout",
    "stderr", "defined", "pragma",
}

IGNORE_CALLER_PREFIXES = (
    "Rte_Runnable_", "Rte_Call_", "Rte_Read_", "Rte_Write_",
    "Rte_Mode_", "Rte_Irv_", "Rte_Enter_", "Rte_Exit_",
)

ROOT_NAME_PREFIXES = ("OsTask_",)
ROOT_NAME_EXACT = {"main", "Main", "Reset_Handler", "Default_Handler"}

SKIP_DIR_NAMES = {
    ".git", ".svn", ".cursor", "venv", "node_modules",
    "__pycache__", ".idea", ".vs", "build", "out",
}

MAX_CHAIN_DEPTH = 40
CALL_PATTERN = re.compile(r"\b([A-Za-z_]\w*)\s*\(")

# Function header patterns (opening brace may be on same line or next non-empty line)
FUNC_HEADER_RE = re.compile(
    r"^\s*(?:static\s+)?(?:inline\s+)?(?:const\s+)?[\w\s\*]+?\s+([A-Za-z_]\w*)\s*\([^;]*\)\s*$"
)
AUTOSAR_FUNC_HEADER_RE = re.compile(
    r"^\s*FUNC\s*\([^)]+\)\s*([A-Za-z_]\w*)\s*\([^;]*\)\s*$"
)
TASK_HEADER_RE = re.compile(r"^\s*TASK\s*\(\s*([A-Za-z_]\w*)\s*\)\s*$")

# Same-line brace variants
FUNC_INLINE_RE = re.compile(
    r"^\s*(?:static\s+)?(?:inline\s+)?(?:const\s+)?[\w\s\*]+?\s+([A-Za-z_]\w*)\s*\([^;]*\)\s*\{"
)
AUTOSAR_FUNC_INLINE_RE = re.compile(
    r"^\s*FUNC\s*\([^)]+\)\s*([A-Za-z_]\w*)\s*\([^;]*\)\s*\{"
)
TASK_INLINE_RE = re.compile(r"^\s*TASK\s*\(\s*([A-Za-z_]\w*)\s*\)\s*\{")
AUTOSAR_FUNC_ONLY_RE = re.compile(r"^\s*FUNC\s*\([^)]+\)\s*$")


@dataclass(frozen=True)
class FunctionDef:
    name: str
    file_path: Path
    is_static: bool
    line_number: int


@dataclass
class CallerRef:
    function: str
    file_path: Path


def strip_c_comments(source: str) -> str:
    result: List[str] = []
    i = 0
    n = len(source)
    while i < n:
        if source[i : i + 2] == "//":
            while i < n and source[i] != "\n":
                i += 1
        elif source[i : i + 2] == "/*":
            i += 2
            while i < n and source[i : i + 2] != "*/":
                if source[i] == "\n":
                    result.append("\n")
                i += 1
            i += 2
        else:
            result.append(source[i])
            i += 1
    return "".join(result)


def count_braces_in_line(line: str) -> int:
    """Return net brace depth change, ignoring braces inside string literals."""
    delta = 0
    in_string = False
    escape = False
    quote = ""
    for ch in line:
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == quote:
                in_string = False
            continue
        if ch in ("'", '"'):
            in_string = True
            quote = ch
        elif ch == "{":
            delta += 1
        elif ch == "}":
            delta -= 1
    return delta


def is_skippable_line(stripped: str) -> bool:
    return (
        not stripped
        or stripped.startswith("//")
        or stripped.startswith("/*")
        or stripped.startswith("*")
        or stripped == "*/"
    )


def extract_name_before_paren(signature: str) -> Optional[str]:
    """Extract function name from a C signature fragment (without '{')."""
    m = re.search(r"\b([A-Za-z_]\w*)\s*\(", signature)
    if m and m.group(1) not in C_KEYWORDS:
        return m.group(1)
    return None


def parse_function_header(line: str) -> Optional[Tuple[str, bool]]:
    """Return (name, is_static) if line starts a function, else None."""
    if "(" not in line or ")" not in line:
        return None

    for pattern in (
        AUTOSAR_FUNC_INLINE_RE,
        TASK_INLINE_RE,
        FUNC_INLINE_RE,
        AUTOSAR_FUNC_HEADER_RE,
        TASK_HEADER_RE,
        FUNC_HEADER_RE,
    ):
        m = pattern.match(line)
        if not m:
            continue
        name = m.group(1)
        if name in C_KEYWORDS:
            return None
        return name, bool(re.search(r"\bstatic\b", line))
    return None


def discover_functions(source: str, file_path: Path) -> List[Tuple[FunctionDef, str]]:
    """
    Single-pass line scanner — O(lines) per file, no repeated brace matching.
    Returns (FunctionDef, body_text) for each function found.
    """
    lines = source.splitlines()
    functions: List[Tuple[FunctionDef, str]] = []

    depth = 0
    cur_name: Optional[str] = None
    cur_static = False
    cur_start_line = 0
    body_lines: List[str] = []
    pending_header: Optional[Tuple[str, bool, int]] = None
    after_autosar_func = False
    autosar_pending_name: Optional[str] = None
    autosar_start_line = 0

    def start_function(name: str, is_static: bool, start_line: int, open_line: str) -> None:
        nonlocal depth, cur_name, cur_static, cur_start_line, body_lines
        nonlocal after_autosar_func, autosar_pending_name, pending_header
        cur_name = name
        cur_static = is_static
        cur_start_line = start_line
        body_lines = []
        depth = count_braces_in_line(open_line)
        after_autosar_func = False
        autosar_pending_name = None
        pending_header = None
        if depth <= 0:
            functions.append((FunctionDef(name, file_path, is_static, start_line), ""))
            cur_name = None
            body_lines = []
            depth = 0

    for lineno, line in enumerate(lines, start=1):
        stripped = line.strip()

        if depth == 0:
            header = parse_function_header(line)
            if header:
                name, is_static = header
                if "{" in line:
                    start_function(name, is_static, lineno, line)
                else:
                    pending_header = (name, is_static, lineno)
                    after_autosar_func = False
                    autosar_pending_name = None
                continue

            if AUTOSAR_FUNC_ONLY_RE.match(stripped):
                after_autosar_func = True
                autosar_pending_name = None
                autosar_start_line = lineno
                pending_header = None
                continue

            if after_autosar_func:
                if is_skippable_line(stripped):
                    continue

                if "{" in line:
                    name = autosar_pending_name
                    if name is None:
                        name = extract_name_before_paren(line.split("{", 1)[0])
                    if name and name not in C_KEYWORDS:
                        start_function(name, False, autosar_start_line, line)
                    else:
                        after_autosar_func = False
                        autosar_pending_name = None
                    continue

                if autosar_pending_name is None:
                    if "(" in line:
                        name = extract_name_before_paren(line)
                        if name:
                            autosar_pending_name = name
                            autosar_start_line = lineno
                    else:
                        m = re.match(r"^([A-Za-z_]\w*)", stripped)
                        if m and m.group(1) not in C_KEYWORDS:
                            autosar_pending_name = m.group(1)
                            autosar_start_line = lineno
                continue

            if pending_header and stripped == "{":
                name, is_static, start_line = pending_header
                start_function(name, is_static, start_line, line)
                continue

            pending_header = None
            continue

        # Inside a function body
        body_lines.append(line)
        depth += count_braces_in_line(line)

        if depth <= 0:
            functions.append(
                (
                    FunctionDef(cur_name or "<unknown>", file_path, cur_static, cur_start_line),
                    "\n".join(body_lines),
                )
            )
            cur_name = None
            body_lines = []
            depth = 0
            pending_header = None

    unique: Dict[str, Tuple[FunctionDef, str]] = {}
    for func, body in functions:
        if func.name not in unique:
            unique[func.name] = (func, body)
    return list(unique.values())


def extract_calls(body: str) -> Set[str]:
    return {
        m.group(1)
        for m in CALL_PATTERN.finditer(body)
        if m.group(1) not in C_KEYWORDS
    }


def is_root_function(name: str) -> bool:
    if name in ROOT_NAME_EXACT:
        return True
    return any(name.startswith(p) for p in ROOT_NAME_PREFIXES)


def should_skip_caller(name: str) -> bool:
    return any(name.startswith(p) for p in IGNORE_CALLER_PREFIXES)


def should_skip_dir(path: Path) -> bool:
    return any(part in SKIP_DIR_NAMES for part in path.parts)


class CallGraphIndex:
    def __init__(self) -> None:
        self.functions: Dict[str, List[FunctionDef]] = defaultdict(list)
        self.callers: Dict[str, List[CallerRef]] = defaultdict(list)
        self._file_cache: Dict[Path, str] = {}

    def scan_directory(self, src_dir: Path, extensions: Tuple[str, ...] = (".c",)) -> int:
        files_scanned = 0
        for path in sorted(src_dir.rglob("*")):
            if not path.is_file() or should_skip_dir(path):
                continue
            if path.suffix.lower() not in extensions:
                continue
            self._index_file(path)
            files_scanned += 1
        return files_scanned

    def _read_file(self, path: Path) -> str:
        if path not in self._file_cache:
            self._file_cache[path] = strip_c_comments(
                path.read_text(encoding="utf-8", errors="replace")
            )
        return self._file_cache[path]

    def _index_file(self, path: Path) -> None:
        for func, body in discover_functions(self._read_file(path), path):
            self.functions[func.name].append(func)
            for callee in extract_calls(body):
                self.callers[callee].append(CallerRef(func.name, path))

    def resolve_callers(self, func_name: str, defined_in: Optional[Path] = None) -> List[CallerRef]:
        filtered: List[CallerRef] = []
        seen: Set[Tuple[str, str]] = set()

        for ref in self.callers.get(func_name, []):
            if should_skip_caller(ref.function):
                continue
            key = (ref.function, str(ref.file_path))
            if key in seen:
                continue
            seen.add(key)

            if defined_in is not None:
                defs = [d for d in self.functions.get(func_name, []) if d.file_path == defined_in]
                if defs and defs[0].is_static and ref.file_path != defined_in:
                    continue

            filtered.append(ref)

        return filtered

    @staticmethod
    def _caller_rank(ref: CallerRef) -> Tuple[int, int, str]:
        score = 0
        if is_root_function(ref.function):
            score = 4
        elif ref.function.startswith("RE_"):
            score = 3
        elif ref.function.endswith("_step") or ref.function.endswith("_initialize"):
            score = 2
        elif ref.function.startswith("ICU_") or ref.function.startswith("AppMode_"):
            score = 1
        rte_penalty = 0 if ref.file_path.name == "Rte.c" else 1
        return (-score, -rte_penalty, ref.function)

    def pick_best_chain(self, func_name: str, defined_in: Path) -> List[CallerRef]:
        """Greedy upward walk — bounded depth, no exponential DFS."""
        chain: List[CallerRef] = []
        current = func_name
        current_file: Optional[Path] = defined_in
        visited: Set[str] = set()

        for _ in range(MAX_CHAIN_DEPTH):
            if current in visited:
                break
            visited.add(current)
            if is_root_function(current):
                break

            callers = self.resolve_callers(current, current_file)
            if not callers:
                break

            best = min(callers, key=self._caller_rank)
            chain.append(best)
            current = best.function
            current_file = best.file_path

        return chain


@dataclass
class BacktraceEntry:
    target_file: str
    target_function: str
    chain: List[CallerRef]


def find_c_file(src_dir: Path, filename: str) -> Path:
    matches = [p for p in src_dir.rglob(filename) if p.is_file() and not should_skip_dir(p)]
    if not matches:
        raise FileNotFoundError(f"No file named '{filename}' found under {src_dir}")
    if len(matches) > 1:
        matches.sort(key=lambda p: (len(p.parts), str(p)))
        print(f"Warning: multiple '{filename}' found; using {matches[0]}", file=sys.stderr)
    return matches[0]


def backtrace_file(index: CallGraphIndex, target_path: Path) -> List[BacktraceEntry]:
    source = strip_c_comments(target_path.read_text(encoding="utf-8", errors="replace"))
    return [
        BacktraceEntry(
            target_file=target_path.name,
            target_function=func.name,
            chain=index.pick_best_chain(func.name, target_path),
        )
        for func, _ in discover_functions(source, target_path)
    ]


def write_excel(
    entries: List[BacktraceEntry],
    output_path: Path,
    sheet_name: str,
    src_dir: Optional[Path] = None,
    target_file: Optional[str] = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Record scan inputs at top of sheet
    if src_dir is not None:
        ws.cell(row=1, column=3, value="Source directory")
        ws.cell(row=1, column=5, value=str(src_dir))
    if target_file is not None:
        ws.cell(row=2, column=3, value="Input .c file")
        ws.cell(row=2, column=5, value=target_file)

    header_row = 4
    ws.cell(row=header_row, column=3, value="Target file")
    ws.cell(row=header_row, column=5, value="Target function")
    ws.cell(row=header_row, column=7, value="Interface function")
    ws.cell(row=header_row, column=9, value="Interface file")

    row = header_row + 2
    for entry in entries:
        if not entry.chain:
            ws.cell(row=row, column=3, value=entry.target_file)
            ws.cell(row=row, column=5, value=entry.target_function)
            ws.cell(row=row, column=7, value="(no caller found)")
            row += 2
            continue

        first = entry.chain[0]
        ws.cell(row=row, column=3, value=entry.target_file)
        ws.cell(row=row, column=5, value=entry.target_function)
        ws.cell(row=row, column=7, value=first.function)
        ws.cell(row=row, column=9, value=first.file_path.name)
        row += 1
        for caller in entry.chain[1:]:
            ws.cell(row=row, column=7, value=caller.function)
            ws.cell(row=row, column=9, value=caller.file_path.name)
            row += 1
        row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_summary(entries: List[BacktraceEntry]) -> None:
    for entry in entries:
        print(f"\nTARGET: {entry.target_function} in {entry.target_file}")
        if not entry.chain:
            print("  (no caller found in scanned source tree)")
            continue
        for caller in entry.chain:
            print(f"  <- {caller.function} ({caller.file_path.name})")


def prompt_inputs() -> Tuple[Path, str, Optional[Path]]:
    """Ask for source directory and .c file name when CLI args are not given."""
    print("C Function Backtracer — enter inputs\n")

    default_src = Path(__file__).resolve().parent.parent
    src_raw = input(f"Source code directory [{default_src}]: ").strip()
    src_dir = Path(src_raw or default_src).resolve()

    while True:
        file_name = input("Target .c file name (e.g. PDC_FrunkLamp.c): ").strip()
        if file_name:
            if not file_name.lower().endswith(".c"):
                file_name += ".c"
            break
        print("  Please enter a .c file name.")

    out_raw = input("Output Excel path (leave blank for auto): ").strip()
    output = Path(out_raw).resolve() if out_raw else None

    return src_dir, file_name, output


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backtrace all functions in a C file and export caller chains to Excel.",
        epilog="Run without --src-dir and --file to be prompted for inputs interactively.",
    )
    parser.add_argument("--src-dir", type=Path, default=None,
                        help="Root directory to scan for C source files (recursive).")
    parser.add_argument("--file", default=None,
                        help="Target .c file name (e.g. PDC_FrunkLamp.c).")
    parser.add_argument("-o", "--output", type=Path, default=None,
                        help="Output .xlsx path (default: <file_stem>_backtrace.xlsx).")
    parser.add_argument("--no-excel", action="store_true",
                        help="Print summary only; skip Excel output.")
    parser.add_argument("--extensions", default=".c",
                        help="Comma-separated extensions to scan (default: .c).")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    if args.src_dir is None or args.file is None:
        src_dir, file_name, prompted_output = prompt_inputs()
        if args.output is None:
            args.output = prompted_output
    else:
        src_dir = args.src_dir.resolve()
        file_name = args.file
        if not file_name.lower().endswith(".c"):
            file_name += ".c"

    args.file = file_name
    if not src_dir.is_dir():
        print(f"Error: --src-dir is not a directory: {src_dir}", file=sys.stderr)
        return 1

    try:
        target_path = find_c_file(src_dir, args.file)
    except FileNotFoundError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    extensions = tuple(
        e if e.startswith(".") else f".{e}"
        for part in args.extensions.split(",")
        if (e := part.strip())
    )

    print(f"Scanning: {src_dir}", flush=True)
    index = CallGraphIndex()
    count = index.scan_directory(src_dir, extensions)
    print(f"Indexed {count} file(s).", flush=True)

    print(f"Backtracing: {target_path}", flush=True)
    entries = backtrace_file(index, target_path)
    print(f"Found {len(entries)} function(s).", flush=True)

    print_summary(entries)

    if not args.no_excel:
        output = (args.output or Path(f"{target_path.stem}_backtrace.xlsx")).resolve()
        write_excel(
            entries, output, target_path.stem[:31],
            src_dir=src_dir, target_file=args.file,
        )
        print(f"\nExcel written to: {output}", flush=True)

    return 0


if __name__ == "__main__":
    sys.exit(main())
