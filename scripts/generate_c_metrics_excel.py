"""
Generate Project_report.xlsx from a source-code tree.

Logic taken verbatim from "script for statement and no.of lines.txt":
  - count ALL lines (including blank & comment lines)
  - count statements as the number of ';' characters per line

Two sheets are produced:

  Sheet 1 — "Module Summary"
    S No | Module No. | Folder Path | C File | No. of Lines | Statements |
    Executable SLOC
    (each .c file is on its own row; Module No. & Folder only on the first
     row of each module — the remaining rows in that module leave them blank)

  Sheet 2 — "report"
    S NO | UT | IT | Path | C File | No. of Lines | Statements |
    Compilation | Actual Sloc | Remarks | Engineer Name

ROOT_PATH and OUTPUT_DIR are injected by the GUI wrapper.
"""

import os
import sys

from stdio_utils import ensure_utf8_stdio

ensure_utf8_stdio()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

ROOT_PATH      = r"D:/Source_code/BC4i_P_E2.0_B2412/B2412/Static_Code/KSC"
OUTPUT_DIR     = ""
OUTPUT_FILENAME = "Project_report.xlsx"

# ── report (detail) column layout ─────────────────────────────────────────────
DETAIL_HEADERS = [
    "S NO", "UT", "IT", "Path", "C File",
    "No. of Lines", "Statements", "Compilation", "Actual Sloc", "Remarks",
    "Engineer Name",
]
DETAIL_COL_WIDTHS = {
    "A":  5.5, "B":  3.5, "C":  3.0,
    "D": 80.0, "E": 47.0,
    "F": 12.0, "G": 12.0,
    "H": 20.0, "I": 11.0, "J": 19.0, "K": 19.0,
}

# ── Module Summary column layout ───────────────────────────────────────────────
SUMMARY_HEADERS = [
    "S No", "Module No.", "Folder Path", "C File", "No. of Lines", "Statements",
    "Executable SLOC",
]
SUMMARY_COL_WIDTHS = {
    "A":  5.5, "B": 10.0, "C": 80.0, "D": 45.0, "E": 14.0, "F": 13.0, "G": 16.0,
}


# ── core analysis (identical to the reference script) ─────────────────────────
def analyze_c_file(file_path: str) -> tuple[int, int]:
    """Return (total_lines, statements_count) — counts ALL lines & every ';'."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as fh:
            lines = fh.readlines()
        line_count       = len(lines)                             # all lines
        statements_count = sum(line.count(";") for line in lines) # ';' count
        return line_count, statements_count
    except OSError as exc:
        print(f"[WARN] Error reading {file_path}: {exc}")
        return 0, 0


def _scan_tree(root_path: str) -> list[tuple[str, str, int, int]]:
    """
    Walk every folder under root_path, collect all .c files.
    Returns sorted list of (folder, c_file, line_count, stmt_count).
    """
    rows: list[tuple[str, str, int, int]] = []
    root_path = os.path.normpath(root_path)

    for folder, _dirs, files in os.walk(root_path):
        c_files = sorted(f for f in files if f.endswith(".c"))
        for c_file in c_files:
            file_path = os.path.join(folder, c_file)
            line_count, stmt_count = analyze_c_file(file_path)
            rows.append((folder, c_file, line_count, stmt_count))
            print(f"[SCAN] {c_file}  ({line_count} lines, {stmt_count} statements)")

    # sort by folder then filename (same ordering as reference script)
    rows.sort(key=lambda r: (r[0].lower(), r[1].lower()))
    return rows


def _assign_it_numbers(
    root_path: str, rows: list[tuple[str, str, int, int]]
) -> list[int | None]:
    """
    IT numbering that matches Project_report.xlsx:
      - Files directly in root  → every .c file gets its own IT number
      - Files in a sub-folder   → first .c file in each folder gets the
                                   next IT number; siblings leave IT blank
    """
    root_path = os.path.normpath(root_path)
    it_values: list[int | None] = []
    it_counter  = 0
    seen_folders: set[str] = set()

    for folder, _c_file, _lines, _stmts in rows:
        folder = os.path.normpath(folder)
        if folder == root_path:                   # root-level file
            it_counter += 1
            it_values.append(it_counter)
        elif folder not in seen_folders:          # first file of a sub-folder
            seen_folders.add(folder)
            it_counter += 1
            it_values.append(it_counter)
        else:                                     # subsequent files in same sub-folder
            it_values.append(None)

    return it_values


# ── styling helpers ───────────────────────────────────────────────────────────
def _thin_border() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, headers: list[str], col_widths: dict[str, float]) -> None:
    hdr_font  = Font(bold=True, size=11)
    hdr_fill  = PatternFill("solid", fgColor="FFFF00")   # yellow — matches sample
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border    = _thin_border()

    for col_idx, title in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes       = "A2"
    ws.row_dimensions[1].height = 30


# ── Module Summary sheet ──────────────────────────────────────────────────────
def _write_summary_sheet(wb: "Workbook", root_path: str, rows: list[tuple]) -> None:
    """
    Each .c file gets its OWN row.
    Module No. and Folder Path appear only on the FIRST row of each module;
    subsequent files in the same folder leave those two columns blank.
    """
    ws = wb.create_sheet("Module Summary", 0)
    _apply_header(ws, SUMMARY_HEADERS, SUMMARY_COL_WIDTHS)

    border      = _thin_border()
    center      = Alignment(horizontal="center", vertical="center")
    left_mid    = Alignment(horizontal="left",   vertical="center")

    root_path   = os.path.normpath(root_path)
    mod_counter = 0
    seen_folders: set[str] = set()
    excel_row   = 2          # row 1 is the header
    s_no        = 0          # continuous serial number

    for folder, c_file, line_count, stmt_count in rows:
        folder = os.path.normpath(folder)
        s_no += 1

        # Module No. — only on the FIRST .c file of each folder
        # Folder Path — ALWAYS written on every row (repeated for the whole module)
        if folder not in seen_folders:
            seen_folders.add(folder)
            mod_counter += 1
            mod_no_val = mod_counter
        else:
            mod_no_val = None   # blank for subsequent files in same folder

        # S No is always filled continuously; Executable SLOC filled after UT compile
        values = [s_no, mod_no_val, folder, c_file, line_count, stmt_count, None]

        for col_idx, val in enumerate(values, start=1):
            cell           = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.border    = border
            # cols 1(S No), 2(Mod No), 5(Lines), 6(Stmts), 7(Exec SLOC) → center
            cell.alignment = center if col_idx in (1, 2, 5, 6, 7) else left_mid

        excel_row += 1

    print(f"[SHEET] 'Module Summary' written — {mod_counter} module(s), {len(rows)} file row(s)")


# ── report (detail) sheet ─────────────────────────────────────────────────────
def _write_detail_sheet(wb: "Workbook", root_path: str, rows: list[tuple]) -> None:
    """Write the report detail sheet."""
    ws = wb.create_sheet("report")
    _apply_header(ws, DETAIL_HEADERS, DETAIL_COL_WIDTHS)

    it_numbers = _assign_it_numbers(root_path, rows)
    border     = _thin_border()
    center     = Alignment(horizontal="center", vertical="center")
    left_mid   = Alignment(horizontal="left",   vertical="center")

    for idx, ((folder, c_file, line_count, stmt_count), it_no) in enumerate(
        zip(rows, it_numbers), start=1
    ):
        row_idx = idx + 1
        values  = [
            idx,        # S NO
            None,       # UT          (manual entry)
            it_no,      # IT          (module number)
            folder,     # Path
            c_file,     # C File
            line_count, # No. of Lines
            stmt_count, # Statements
            None,       # Compilation (manual entry)
            None,       # Actual Sloc (manual entry)
            None,       # Remarks       (manual entry)
            None,       # Engineer Name (manual entry)
        ]
        for col_idx, val in enumerate(values, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = border
            # numeric / short columns → center; path / filename → left
            cell.alignment = center if col_idx in (1, 2, 3, 6, 7, 8, 9) else left_mid

    print(f"[SHEET] 'report' written — {len(rows)} file row(s)")


# ── entry point ───────────────────────────────────────────────────────────────
def main() -> int:
    if Workbook is None:
        print("[ERROR] openpyxl is not installed.  Run: pip install openpyxl")
        return 1

    root_path = (ROOT_PATH or "").strip()
    if not root_path:
        print("[ERROR] Source root path is required.")
        return 1
    if not os.path.isdir(root_path):
        print(f"[ERROR] Source root path does not exist: {root_path}")
        return 1

    out_dir  = (OUTPUT_DIR or "").strip() or root_path
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, OUTPUT_FILENAME)

    print("=" * 55)
    print("  Project Report Excel Generator")
    print("=" * 55)
    print(f"[INIT] Source root : {root_path}")
    print(f"[INIT] Output dir  : {out_dir}")
    print(f"[INIT] Output file : {out_file}")
    print()

    rows = _scan_tree(root_path)
    if not rows:
        print("[WARN] No .c files found under the source root path.")
        return 1

    # Build workbook — two sheets
    wb = Workbook()
    if wb.active and wb.active.title == "Sheet":
        del wb["Sheet"]

    _write_summary_sheet(wb, root_path, rows)
    _write_detail_sheet (wb, root_path, rows)
    wb.save(out_file)

    print()
    print(f"[SUCCESS] .c files scanned   : {len(rows)}")
    print(f"[SUCCESS] Excel saved        : {out_file}")
    print(f"[SUCCESS] Sheets             : Module Summary  +  report")
    return 0


if __name__ == "__main__":
    sys.exit(main())
