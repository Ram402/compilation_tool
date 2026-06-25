"""
VectorCAST Batch Compilation Runner
====================================
Compiles every module listed in MODULES using the exact same logic as
vcast_auto_compile3.py — without modifying that script at all.

How it works
------------
1.  For each module the runner searches HEADER_SEARCH_ROOT for a .c file
    whose stem matches the UUT name.
2.  It derives SOURCE_DIR from the directory that contains the found .c file.
3.  It overrides all the global variables in vcast_auto_compile3 to point at
    the current module, then calls compile_one_module() which runs the full
    auto-retry loop (header fix + macro fix + linker stub).
4.  If a module fails for any reason the exception is caught, the failure is
    recorded, and compilation continues for the next module.
5.  A final summary table is printed showing PASS / FAIL for every module.

Configuration — edit the section below
---------------------------------------
"""

import os
import sys
import re
import subprocess
import platform

from stdio_utils import ensure_utf8_stdio
from stop_control import check as check_stop

ensure_utf8_stdio()

import importlib.util
import traceback
from datetime import datetime
from pathlib import Path

# ============================================================================
# BATCH CONFIGURATION  –  edit these
# ============================================================================

VECTORCAST_DIR     = r"C:\VCAST"
BASE_DIR_NAME      = "R2"
BASE_DIR_PATH      = r"D:\project_4\BC4i_P E2.0_B2412\B2412"
HEADER_SEARCH_ROOT = BASE_DIR_PATH
WORKSPACE_ROOT     = r"D:\Nagesh\workspace\UT"   # each module gets its own subfolder here
MAX_RETRY_ROUNDS   = 100
STOP_FILE          = ""   # set by GUI wrapper; polled for Stop button
IMPORTED_EXCEL_PATH = ""  # set by GUI wrapper if Excel imported

# The original compilation script (keep it next to this file, or give full path)
COMPILE_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "vcast_auto_compile3.py")

# -----------------------------------------------------------------------
# MODULES TABLE
# Each entry:  (UUT_stem, .c_filename)
#   UUT_stem   – used for ENV_NAME, WORK_DIR subfolder, ENVIRO.UUT
#   .c_filename – the actual source file to find under HEADER_SEARCH_ROOT
# -----------------------------------------------------------------------
MODULES = [

    ("Aswc_Brake_Lp",                       "Aswc_Brake_Lp.c"),
    ("PDC_BrakeLamp",                  "PDC_BrakeLamp.c"),

]
# ============================================================================
# BATCH LOG
# ============================================================================
BATCH_LOG = os.path.join(WORKSPACE_ROOT, "batch_compile_log.txt")


def blog(msg: str, also_print: bool = True) -> None:
    os.makedirs(WORKSPACE_ROOT, exist_ok=True)
    with open(BATCH_LOG, "a", encoding="utf-8") as f:
        f.write(msg + "\n")
    if also_print:
        print(msg)


# ============================================================================
# PARSING AND EXCEL UPDATE HELPERS
# ============================================================================

def parse_sloc_from_html(html_path: str):
    import re
    try:
        if not os.path.exists(html_path):
            return None
        with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        # Find header to locate the "Statements" column index
        header_match = re.search(r"<thead[^>]*>.*?</tr>", content, re.DOTALL | re.IGNORECASE)
        if not header_match:
            header_match = re.search(r"<tr>\s*<(?:td|th)[^>]*>Unit.*?\s*</tr>", content, re.DOTALL | re.IGNORECASE)

        statements_idx = 3 # default fallback
        if header_match:
            header_cells = re.findall(r"<(?:td|th)[^>]*>(.*?)</(?:td|th)>", header_match.group(0), re.DOTALL | re.IGNORECASE)
            header_cells = [re.sub(r"<[^>]+>", "", c).strip().lower() for c in header_cells]
            if "statements" in header_cells:
                statements_idx = header_cells.index("statements")

        # Find Grand Totals or Totals row
        match = re.search(r"<tr[^>]*>[^<]*<(?:td|th)[^>]*>\s*GRAND TOTALS.*?</tr>", content, re.DOTALL | re.IGNORECASE)
        if not match:
            match = re.search(r"<tr[^>]*>[^<]*<(?:td|th)[^>]*>\s*TOTALS.*?</tr>", content, re.DOTALL | re.IGNORECASE)

        if match:
            row_content = match.group(0)
            cells = re.findall(r"<(?:td|th)[^>]*>(.*?)</(?:td|th)>", row_content, re.DOTALL | re.IGNORECASE)
            if len(cells) > statements_idx:
                statements_str = re.sub(r"<[^>]+>", "", cells[statements_idx].strip())
                slash_match = re.search(r"/\s*(\d+)", statements_str)
                if slash_match:
                    return int(slash_match.group(1))
                else:
                    num_match = re.search(r"(\d+)", statements_str)
                    if num_match:
                        return int(num_match.group(1))
    except Exception as e:
        blog(f"  [ERROR] Parsing HTML metrics {html_path}: {e}")
    return None


def update_excel_sloc(excel_path: str, c_filename: str, sloc: int) -> tuple:
    if not excel_path or not os.path.exists(excel_path):
        return False, f"Excel path '{excel_path}' does not exist."
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)

        sloc_names = ("executable sloc", "actual sloc", "actual_sloc")
        sheet_names = []
        if "Module Summary" in wb.sheetnames:
            sheet_names.append("Module Summary")
        sheet_names.extend(n for n in wb.sheetnames if n not in sheet_names)

        updated_sheets: list[str] = []
        c_target = c_filename.strip().lower()

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            header_row = 1
            headers: list[str] = []

            for r in range(1, min(10, ws.max_row + 1)):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                row_vals_lower = [
                    str(v).strip().lower() if v is not None else "" for v in row_vals
                ]
                if "c file" in row_vals_lower:
                    header_row = r
                    headers = row_vals_lower
                    break

            if not headers:
                continue

            c_file_col = None
            sloc_col = None
            for idx, h in enumerate(headers, 1):
                if h == "c file":
                    c_file_col = idx
                elif h in sloc_names and sloc_col is None:
                    sloc_col = idx

            if not c_file_col or not sloc_col:
                continue

            target_row = None
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row=r, column=c_file_col).value
                if val is None:
                    continue
                val_str = str(val).strip().lower()
                if val_str == c_target or os.path.basename(val_str) == c_target:
                    target_row = r
                    break

            if target_row:
                ws.cell(row=target_row, column=sloc_col, value=sloc)
                updated_sheets.append(sheet_name)

        if not updated_sheets:
            return False, f"Could not find row for file '{c_filename}' in Excel."

        wb.save(excel_path)
        sheets = ", ".join(updated_sheets)
        return True, f"Executable SLOC updated for {c_filename} to {sloc} ({sheets})"
    except PermissionError:
        return False, f"Permission denied when saving '{excel_path}'. Please close the Excel file."
    except Exception as e:
        return False, f"Error updating Excel: {e}"


# ============================================================================
# FIND .c FILE ON DISK
# ============================================================================

def find_c_file(filename: str, root: str) -> str:
    """
    Case-insensitive walk of root to find filename.
    Returns the full path of the first match, or '' if not found.
    """
    target = filename.lower()
    for dirpath, _dirs, files in os.walk(root):
        for fn in files:
            if fn.lower() == target:
                return os.path.join(dirpath, fn)
    return ""


# ============================================================================
# LOAD THE ORIGINAL COMPILE SCRIPT AS A MODULE
# ============================================================================

def _load_compile_module():
    """
    Dynamically load vcast_auto_compile3.py as a Python module object.
    We load it once and re-use it, overriding its globals per module.
    """
    spec = importlib.util.spec_from_file_location("vcast_compile", COMPILE_SCRIPT)
    mod  = importlib.util.module_from_spec(spec)
    # Execute the module body (defines all functions, sets initial globals)
    spec.loader.exec_module(mod)
    return mod


# ============================================================================
# COMPILE ONE MODULE
# ============================================================================

def compile_one_module(mod, uut_stem: str, c_file: str, source_dir: str) -> bool:
    """
    Override the globals in *mod* to point at this module, then run main().

    Returns True on success, False on failure.
    """
    work_dir   = os.path.join(WORKSPACE_ROOT, uut_stem)
    env_name   = uut_stem
    env_script = os.path.join(work_dir, f"{env_name}.env")
    cfg_file   = os.path.join(work_dir, "CCAST_.CFG")
    build_log  = os.path.join(work_dir, "build_log.txt")
    detailed   = os.path.join(work_dir, "detailed_log.txt")
    error_log  = os.path.join(work_dir, "error_log.txt")

    os.makedirs(work_dir, exist_ok=True)

    # ── Patch ALL relevant globals in the compile module ──────────────
    mod.ENV_NAME           = env_name
    mod.WORK_DIR           = work_dir
    mod.UUT_FILE           = uut_stem
    mod.SOURCE_DIR_1       = source_dir
    mod.SOURCE_DIR_2       = ""
    mod.SOURCE_DIR_3       = ""
    mod.BASE_DIR_NAME      = BASE_DIR_NAME
    mod.BASE_DIR_PATH      = BASE_DIR_PATH
    mod.VECTORCAST_DIR     = VECTORCAST_DIR
    mod.HEADER_SEARCH_ROOT = HEADER_SEARCH_ROOT
    mod.MAX_RETRY_ROUNDS   = MAX_RETRY_ROUNDS
    mod.STOP_FILE          = STOP_FILE
    mod.DEFINES            = list(DEFINES)   # inherit UI-provided defines; reset per-module so auto-fix additions don't bleed across modules
    mod.EXTRA_INCLUDE_1    = ""
    mod.EXTRA_INCLUDE_2    = ""
    mod.EXTRA_INCLUDE_3    = ""

    # Derived paths
    mod.BUILD_LOG    = build_log
    mod.DETAILED_LOG = detailed
    mod.ERROR_LOG    = error_log
    mod.CLICAST_EXE  = os.path.join(VECTORCAST_DIR, "clicast.exe")
    mod.ENV_SCRIPT   = env_script
    mod.CFG_FILE     = cfg_file

    # Override fail() so it raises instead of calling sys.exit()
    # (allows the batch runner to catch the failure and move on)
    def _fail_raise(reason: str):
        raise RuntimeError(f"[FAILED] {reason}")
    mod.fail = _fail_raise

    # Override show_alert() to suppress popups during batch run
    mod.show_alert = lambda title, message, icon="Information": None

    # Override input() prompt so it doesn't block during batch
    import builtins
    _orig_input = builtins.input
    builtins.input = lambda prompt="": None

    try:
        mod.main()
        return True
    except SystemExit as exc:
        if getattr(exc, "code", None) == 2:
            raise
        blog(f"  [FAIL] {uut_stem}: {exc}")
        return False
    except RuntimeError as exc:
        blog(f"  [FAIL] {uut_stem}: {exc}")
        return False
    except Exception as exc:
        blog(f"  [FAIL] {uut_stem}: Unexpected error: {exc}")
        blog(traceback.format_exc(), also_print=False)
        return False
    finally:
        builtins.input = _orig_input


def run_clicast_with_licensing_retry(cmd: list, **kwargs) -> subprocess.CompletedProcess:
    """Run a clicast subprocess, retrying if it encounters a licensing error."""
    max_retries = 10
    retry_delay = 5  # seconds
    
    # Force English locale
    env = kwargs.get("env") or os.environ.copy()
    env["LC_ALL"] = "C"
    env["LANG"] = "C"
    kwargs["env"] = env

    for attempt in range(1, max_retries + 1):
        proc = subprocess.run(cmd, **kwargs)
        
        is_lic = (proc.returncode == 16)
        if not is_lic and proc.returncode != 0:
            stderr_str = proc.stderr or ""
            stdout_str = proc.stdout or ""
            if isinstance(stderr_str, bytes):
                stderr_str = stderr_str.decode("utf-8", errors="replace")
            if isinstance(stdout_str, bytes):
                stdout_str = stdout_str.decode("utf-8", errors="replace")
            if "license" in stderr_str.lower() or "licensing" in stderr_str.lower() or "flexlm" in stderr_str.lower():
                is_lic = True
            elif "license" in stdout_str.lower() or "licensing" in stdout_str.lower() or "flexlm" in stdout_str.lower():
                is_lic = True
                
        if is_lic:
            if attempt < max_retries:
                blog(f"  [LICENSE] Clicast licensing error (exit code {proc.returncode}). Retrying in {retry_delay}s (attempt {attempt}/{max_retries})...")
                import time
                time.sleep(retry_delay)
                continue
            else:
                blog(f"  [LICENSE] Clicast licensing error persisted after {max_retries} attempts.")
        
        return proc


# ============================================================================
# MAIN BATCH RUNNER
# ============================================================================

def main():
    sep  = "=" * 76
    sep2 = "─" * 76

    # Init batch log
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(WORKSPACE_ROOT, exist_ok=True)
    with open(BATCH_LOG, "w", encoding="utf-8") as f:
        f.write(f"{sep}\nVectorCAST Batch Compilation\nStarted: {now}\n{sep}\n\n")
        f.write(f"  Compile script : {COMPILE_SCRIPT}\n")
        f.write(f"  BASE_DIR_PATH  : {BASE_DIR_PATH}\n")
        f.write(f"  WORKSPACE_ROOT : {WORKSPACE_ROOT}\n")
        f.write(f"  Modules        : {len(MODULES)}\n\n")

    print(sep)
    print("  VectorCAST Batch Compilation Runner")
    print(f"  Modules       : {len(MODULES)}")
    print(f"  Base dir      : {BASE_DIR_PATH}")
    print(f"  Workspace     : {WORKSPACE_ROOT}")
    print(f"  Compile script: {COMPILE_SCRIPT}")
    print(sep)

    # Verify compile script exists
    if not os.path.isfile(COMPILE_SCRIPT):
        print(f"\n[ERROR] Compile script not found: {COMPILE_SCRIPT}")
        print("Place vcast_auto_compile3.py in the same folder as this script.")
        sys.exit(1)

    # Load the compile module once
    blog("\n[BATCH] Loading compile module...")
    try:
        mod = _load_compile_module()
        blog(f"  [OK] Loaded: {COMPILE_SCRIPT}")
    except Exception as exc:
        blog(f"[ERROR] Failed to load compile script: {exc}")
        sys.exit(1)

    results = []   # list of (uut_stem, c_file, source_dir, status, duration_s)

    for idx, (uut_stem, c_filename) in enumerate(MODULES, 1):
        check_stop(f"module {idx}/{len(MODULES)}")

        print()
        print(sep2)
        print(f"  MODULE {idx}/{len(MODULES)}: {uut_stem}")
        print(f"  Source file  : {c_filename}")
        print(sep2)
        blog(f"\n[MODULE {idx}/{len(MODULES)}] {uut_stem}  ({c_filename})")

        # ── Step 1: find the .c file ──────────────────────────────────
        blog(f"  Searching for {c_filename} under {HEADER_SEARCH_ROOT} ...")
        c_path = find_c_file(c_filename, HEADER_SEARCH_ROOT)

        if not c_path:
            msg = f"Source file '{c_filename}' not found under '{HEADER_SEARCH_ROOT}'"
            blog(f"  [SKIP] {msg}")
            print(f"\n  [SKIP] {msg}")
            results.append((uut_stem, c_filename, "", "SKIP – file not found", 0))
            continue

        source_dir = os.path.dirname(c_path)
        blog(f"  [FOUND] {c_path}")
        blog(f"  Source dir   : {source_dir}")
        print(f"  Found        : {c_path}")
        print(f"  Source dir   : {source_dir}")

        # ── Step 2: compile ───────────────────────────────────────────
        t_start = datetime.now()
        print(f"\n  Starting compilation at {t_start.strftime('%H:%M:%S')} ...")
        blog(f"  Compiling {uut_stem}...")

        success = compile_one_module(mod, uut_stem, c_filename, source_dir)

        t_end  = datetime.now()
        elapsed = round((t_end - t_start).total_seconds())
        status  = "PASS" if success else "FAIL"

        blog(f"  Result: {status}  ({elapsed}s)")
        print(f"\n  Result: {status}  (elapsed: {elapsed}s)")

        if success:
            # Generate the metrics HTML report inside the environment folder
            work_dir = os.path.join(WORKSPACE_ROOT, uut_stem)
            env_name = uut_stem
            report_name = f"{env_name}.html"
            report_path = os.path.join(work_dir, report_name)

            clicast_exe = os.path.join(VECTORCAST_DIR, "clicast.exe")
            if not os.path.exists(clicast_exe):
                clicast_exe = r"C:\VCAST\clicast.exe"

            cmd = [clicast_exe, "-e", env_name, "REports", "Custom", "MEtrics", report_name]
            blog(f"  [INFO] Generating metrics report: {' '.join(cmd)}")

            old_cwd = os.getcwd()
            os.chdir(work_dir)
            try:
                proc = run_clicast_with_licensing_retry(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    creationflags=subprocess.CREATE_NO_WINDOW if platform.system() == "Windows" else 0
                )
                if proc.returncode != 0:
                    blog(f"  [ERROR] Clicast failed to generate report: {proc.stderr}")
                else:
                    blog(f"  [OK] Metrics report generated: {report_path}")
            except Exception as e:
                blog(f"  [ERROR] Failed to run clicast metrics command: {e}")
            finally:
                os.chdir(old_cwd)

            # Parse report and update imported Excel if path provided
            if os.path.exists(report_path):
                sloc = parse_sloc_from_html(report_path)
                if sloc is not None:
                    blog(f"  [INFO] Parsed statements/SLOC: {sloc}")
                    excel_path = IMPORTED_EXCEL_PATH
                    if excel_path and os.path.exists(excel_path):
                        ok, msg = update_excel_sloc(excel_path, c_filename, sloc)
                        if ok:
                            blog(f"  [SUCCESS] {msg}")
                        else:
                            blog(f"  [ERROR] {msg}")
                    else:
                        blog(f"  [WARN] No imported Excel file path provided or file does not exist.")
                else:
                    blog(f"  [ERROR] Could not parse statement count from metrics HTML report.")
            else:
                blog(f"  [ERROR] Metrics report not found: {report_path}")

        results.append((uut_stem, c_filename, source_dir, status, elapsed))

    # ================================================================
    #  FINAL SUMMARY TABLE
    # ================================================================
    print()
    print(sep)
    print("  BATCH COMPILATION SUMMARY")
    print(sep)
    print(f"  {'MODULE':<35} {'FILE':<35} {'STATUS':<8} {'TIME':>6}")
    print(f"  {'─'*35} {'─'*35} {'─'*8} {'─'*6}")

    passed = failed = skipped = 0
    for uut, cfile, sdir, status, elapsed in results:
        icon  = "✓" if status == "PASS" else ("~" if "SKIP" in status else "✗")
        tstr  = f"{elapsed}s" if elapsed else "-"
        print(f"  {icon} {uut:<34} {cfile:<35} {status:<8} {tstr:>6}")
        if status == "PASS":
            passed += 1
        elif "SKIP" in status:
            skipped += 1
        else:
            failed += 1

    print(f"  {'─'*35} {'─'*35} {'─'*8} {'─'*6}")
    print(f"  Total: {len(results)}   PASS: {passed}   FAIL: {failed}   SKIP: {skipped}")
    print(sep)

    # Write summary to batch log
    completed = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    blog(f"\n{sep}")
    blog("BATCH SUMMARY")
    blog(sep)
    for uut, cfile, sdir, status, elapsed in results:
        tstr = f"{elapsed}s" if elapsed else "-"
        blog(f"  {status:<6} {uut:<35} {cfile}  [{tstr}]")
        if sdir:
            blog(f"         Source: {sdir}", also_print=False)
    blog(f"\nCompleted: {completed}")
    blog(f"PASS: {passed}  FAIL: {failed}  SKIP: {skipped}")
    blog(sep)

    print(f"\n  Batch log : {BATCH_LOG}")
    print(f"  Each module has its own detailed_log.txt in:")
    print(f"    {WORKSPACE_ROOT}\\<MODULE_NAME>\\")

    if failed > 0:
        print(f"\n  [NOTE] {failed} module(s) failed.")
        print("  Check <MODULE_NAME>\\build_log.txt for the specific error.")

    print("\nDone.")
    # Let the GUI decide PASS/FAIL based on subprocess exit code.
    # Match the old Tkinter UI behavior: any failed module => non-zero exit.
    sys.exit(1 if failed > 0 else 0)

if __name__ == "__main__":
    main()
