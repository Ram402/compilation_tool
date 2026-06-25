"""Excel Import Tab"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QScrollArea,
    QPushButton, QLabel, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog,
    QSizePolicy, QMessageBox, QSpinBox, QLineEdit
)
from PySide6.QtCore import Qt

from ui.widgets         import SectionBanner, SectionSep
from ui.animated_button import AnimatedRunButton
from ui.button_styles   import BTN_BROWSE, BTN_BLUE
from ui.style_helpers   import UI_FONT, MONO_FONT
from ui.dark_bg         import apply_dark_animated_bg

import os
import base64
import tempfile

C_GREEN  = "#4ade80"
C_ERRCLR = "#f87171"
C_MUTED  = "#94a3b8"
C_BORDER = "#3d5472"
C_BLUE   = "#38bdf8"

# White checkmark, pre-rendered as a small PNG (anti-aliased, transparent
# background). Qt's QSS "image: url(...)" reliably loads real files but does
# NOT reliably support data:-URI images on every Qt build, so we materialize
# this to a real file on disk once and point the stylesheet at that path.
_CHECK_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAACT0lEQVR4nO2YPU8bQRRF"
    "Z8EpCU0sBMISDUJpgMpSQEqT/0oFTRqKKL+BtEmLFAk7UWRIwZdzKPY98Tye9Y5Z41lF"
    "c6XVyLO73nNn7sw+27msrKysrP9NQAEUqTnmFrACrFR9brU88E1gI3SulQJWpe0D58A1"
    "MAI+A/tyrp2RAjrS9oEHpnUD7Mm6aNdMGPgD4EqA74F/ctxK35lct5qW2MjAHwIDAR17"
    "o69GfgPrcv1ElJJMCdApiuIROHTOfXHOdZ1z4xk8yJFeXmx05B8D2Qe4k/ZE7kkboYrY"
    "VMFr/xDYSb6IPfhhDbzuRj+AXbkv3TY6Z2wU/jvQk/uaRYcGtQnPL6mY2Fj47cbwNKxN"
    "PPh5YtN85D3wDWAzdC4C/h1wudTYaFyAfcp6ZERZn5wD/boHGPgucJEiNgVl/XFT8UA1"
    "0amB/xYJv5jYeABn8uW3PL/S76XvCjjwTTAZGx35UIFmTS12t5F2nbL+UHArrVcGlKWA"
    "AzqEYxMD3zw2AQNvgV8VBizAUE0YeI1NHfzPhY28Z0JH8kQedBeiMCAD4D2wFgGvszcC"
    "PshzptZRUwO6iHeo37u1/5L6BWvhj14F3pjQKO1S7hCzRtVGzK/n/f7rV4c3JjRKPcrF"
    "NsvEOAJ+BBwvBT5gYjvCRB38ckZ+hoke9XEKwS8vNpEmYmYiXWwiTNg4hXac9LGpUsVM"
    "2L9BHqRtH7zKmLDvCV+f5Jo3aWkrZEzsAafAH+Av8BX4aK9prZj8wdMFtkLnWi2mf3IW"
    "rR/5kEj9n01WVlZW1kv1BMlZTtixMkS+AAAAAElFTkSuQmCC"
)


def _materialize_check_icon() -> str:
    """Write the checkmark PNG to a stable temp-file path and return it
    (forward slashes, as Qt stylesheets expect)."""
    path = os.path.join(tempfile.gettempdir(), "uut_excel_tab_check_icon.png")
    try:
        with open(path, "wb") as f:
            f.write(base64.b64decode(_CHECK_PNG_B64))
    except OSError:
        pass
    return path.replace("\\", "/")


_CHECK_ICON_PATH = _materialize_check_icon()

# Shared stylesheet for QCheckBox widgets — bright, high-contrast indicator
# box so checkboxes read clearly against the dark theme (matches the table's
# checkbox styling below).
CHECKBOX_QSS = f"""
    QCheckBox {{ spacing: 8px; }}
    QCheckBox::indicator {{
        width: 18px;
        height: 18px;
        border: 2px solid {C_BLUE};
        border-radius: 4px;
        background: #16202e;
    }}
    QCheckBox::indicator:hover {{
        border-color: #7dd3fc;
    }}
    QCheckBox::indicator:checked {{
        background: #16202e;
        border: 2px solid {C_BLUE};
        image: url({_CHECK_ICON_PATH});
    }}
"""

# Stylesheet for the per-row checkbox column inside the QTableWidget
TABLE_CHECKBOX_QSS = f"""
    QTableWidget::indicator {{
        width: 18px;
        height: 18px;
        border: 2px solid {C_BLUE};
        border-radius: 4px;
        background: #16202e;
    }}
    QTableWidget::indicator:unchecked {{
        background: #16202e;
        border: 2px solid {C_BLUE};
    }}
    QTableWidget::indicator:checked {{
        background: #16202e;
        border: 2px solid {C_BLUE};
        image: url({_CHECK_ICON_PATH});
    }}
"""


class ExcelTab(QWidget):
    IMG = "https://images.unsplash.com/photo-1596003906949-67221c37965c?w=1200&q=70"

    def __init__(self, batch_tab, switch_to_batch):
        super().__init__()
        self._batch_tab       = batch_tab
        self._switch_to_batch = switch_to_batch
        self._data            = []
        self._syncing_select_all = False
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        outer.addWidget(SectionBanner(
            "IMPORT EXCEL",
            "Load Module List From Excel / CSV File · Send to UT Compilation",
            self.IMG
        ))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        outer.addWidget(scroll, 1)

        body = QWidget()
        body.setStyleSheet("background:transparent;")
        scroll.setWidget(body)
        v = QVBoxLayout(body)
        v.setContentsMargins(48, 0, 48, 40)
        v.setSpacing(0)

        # ── File picker ───────────────────────────────────────────────
        v.addWidget(SectionSep("File", color="blue"))

        file_row = QHBoxLayout()
        file_row.setSpacing(8)

        self.path_lbl = QLabel("No file selected")
        self.path_lbl.setFont(UI_FONT(13))
        self.path_lbl.setStyleSheet(
            f"color:{C_MUTED}; background:#2e3f56; border:1px solid {C_BORDER};"
            "border-radius:3px; padding:7px 12px;")
        self.path_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.path_lbl.setFixedHeight(36)
        file_row.addWidget(self.path_lbl)

        browse_btn = QPushButton("Browse...")
        browse_btn.setStyleSheet(BTN_BROWSE)
        browse_btn.setFixedWidth(100)
        browse_btn.setFixedHeight(36)
        browse_btn.setCursor(Qt.PointingHandCursor)
        browse_btn.clicked.connect(self._browse)
        file_row.addWidget(browse_btn)
        v.addLayout(file_row)

        # ── Column number picker ──────────────────────────────────────
        col_row = QHBoxLayout()
        col_row.setContentsMargins(0, 12, 0, 0)
        col_row.setSpacing(10)

        col_lbl = QLabel("Column number containing .c filenames:")
        col_lbl.setFont(UI_FONT(12))
        col_lbl.setStyleSheet(f"color:{C_MUTED}; background:transparent; border:none;")
        col_row.addWidget(col_lbl)

        self.col_spin = QSpinBox()
        self.col_spin.setMinimum(1)
        self.col_spin.setMaximum(50)
        self.col_spin.setValue(1)
        self.col_spin.setFixedWidth(70)
        self.col_spin.setFixedHeight(30)
        self.col_spin.setToolTip("1 = Column A, 2 = Column B, etc.")
        self.col_spin.setStyleSheet(f"""
            QSpinBox {{
                background: #2e3f56;
                border: 1px solid {C_BORDER};
                border-radius: 3px;
                color: {C_BLUE};
                font-family: 'Rajdhani', 'Segoe UI', sans-serif;
                font-size: 13px;
                padding: 3px 6px;
            }}
            QSpinBox:focus {{ border-color: {C_BLUE}; }}
            QSpinBox::up-button, QSpinBox::down-button {{
                width: 16px;
                background: #3d5472;
                border: none;
            }}
        """)
        col_row.addWidget(self.col_spin)

        col_hint = QLabel("  (1 = Col A)  ·  UUT_NAME = stem,  UUT_FILE = filename.c")
        col_hint.setFont(UI_FONT(12))
        col_hint.setStyleSheet("color:#64748b; background:transparent; border:none;")
        col_row.addWidget(col_hint)
        col_row.addStretch()
        v.addLayout(col_row)

        # Reload button — re-reads same file with new column setting
        reload_row = QHBoxLayout()
        reload_row.setContentsMargins(0, 8, 0, 0)
        reload_row.setSpacing(8)
        self.reload_btn = QPushButton("Reload preview")
        self.reload_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: 1px solid {C_BORDER};
                border-radius: 3px;
                color: {C_BLUE};
                font-family: 'JetBrains Mono', monospace;
                font-size: 11px;
                padding: 4px 12px;
                min-height: 26px;
            }}
            QPushButton:hover  {{ background: rgba(0,180,216,0.1); border-color: {C_BLUE}; }}
            QPushButton:disabled {{ color: #475569; border-color: #2e3f56; }}
        """)
        self.reload_btn.setCursor(Qt.PointingHandCursor)
        self.reload_btn.setEnabled(False)
        self.reload_btn.clicked.connect(self._reload)
        reload_row.addWidget(self.reload_btn)
        reload_row.addStretch()
        v.addLayout(reload_row)

        self.status_lbl = QLabel("Ready — browse a file and set the column number.")
        self.status_lbl.setFont(UI_FONT(12))
        self.status_lbl.setStyleSheet(f"color:{C_MUTED}; padding-top:6px; border:none;")
        v.addWidget(self.status_lbl)

        info = QLabel(
            "Supported formats: .xlsx / .xls / .csv   "
            "Select the column that contains the .c filenames. "
            "UUT_NAME is derived as the stem (without .c extension) and "
            "UUT_FILE is the full filename — both populated from the same column. "
            "After importing, use the checkboxes below to choose which modules are "
            "sent to the UT Compilation tab, and the search bar to quickly find a .c file."
        )
        info.setFont(UI_FONT(12))
        info.setWordWrap(True)
        info.setStyleSheet(
            f"color:{C_MUTED}; background:rgba(0,180,216,0.05);"
            f"border:1px solid rgba(0,180,216,0.2); border-radius:3px;"
            "padding:10px 14px; margin:10px 0 16px 0;")
        v.addWidget(info)

        # ── Preview table ─────────────────────────────────────────────
        v.addWidget(SectionSep("Preview", color="blue"))

        # Select-all + search controls
        preview_controls = QHBoxLayout()
        preview_controls.setContentsMargins(0, 0, 0, 8)
        preview_controls.setSpacing(10)

        self.select_all_chk = QCheckBox("Select All")
        self.select_all_chk.setChecked(True)
        self.select_all_chk.setFont(UI_FONT(13))
        self.select_all_chk.setStyleSheet(CHECKBOX_QSS)
        self.select_all_chk.toggled.connect(self._on_select_all_toggled)
        preview_controls.addWidget(self.select_all_chk)

        self.selected_count_lbl = QLabel("")
        self.selected_count_lbl.setFont(UI_FONT(12))
        self.selected_count_lbl.setStyleSheet(f"color:{C_MUTED}; background:transparent; border:none;")
        preview_controls.addWidget(self.selected_count_lbl)

        preview_controls.addStretch()

        search_lbl = QLabel("Search:")
        search_lbl.setFont(UI_FONT(12))
        search_lbl.setStyleSheet(f"color:{C_MUTED}; background:transparent; border:none;")
        preview_controls.addWidget(search_lbl)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Filter by .c filename or module name…")
        self.search_edit.setFixedWidth(260)
        self.search_edit.setFixedHeight(30)
        self.search_edit.setStyleSheet(f"""
            QLineEdit {{
                background: #2e3f56;
                border: 1px solid {C_BORDER};
                border-radius: 3px;
                color: {C_BLUE};
                font-family: 'Rajdhani', 'Segoe UI', sans-serif;
                font-size: 13px;
                padding: 3px 6px;
            }}
            QLineEdit:focus {{ border-color: {C_BLUE}; }}
        """)
        self.search_edit.textChanged.connect(self._filter_table)
        preview_controls.addWidget(self.search_edit)

        v.addLayout(preview_controls)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["", "#", "UUT_NAME", "UUT_FILE"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 42)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.table.setColumnWidth(1, 50)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setMinimumHeight(260)
        self.table.setStyleSheet(TABLE_CHECKBOX_QSS)
        self.table.itemChanged.connect(self._on_item_changed)
        v.addWidget(self.table)

        # ── Send to Batch ─────────────────────────────────────────────
        v.addWidget(SectionSep("Send to UT Compilation", color="blue"))
        self.replace_chk = QCheckBox("Replace existing modules (un-check to append)")
        self.replace_chk.setChecked(True)
        self.replace_chk.setFont(UI_FONT(13))
        self.replace_chk.setStyleSheet(CHECKBOX_QSS)
        v.addWidget(self.replace_chk)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 16, 0, 0)
        btn_row.setSpacing(12)

        self.send_btn = AnimatedRunButton("  Send Selected to UT Compilation", color="blue")
        self.send_btn.clicked.connect(self._send_to_batch)
        btn_row.addWidget(self.send_btn)
        btn_row.addStretch()
        v.addLayout(btn_row)

        # Reload preview when the file or column setting changes
        self.col_spin.valueChanged.connect(self._reload)

        v.addStretch()

        # Apply dark gradient + neural field animation background
        apply_dark_animated_bg(self)

    # ── file browse ───────────────────────────────────────────────────
    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel / CSV",
            filter="Spreadsheets (*.xlsx *.xls *.csv);;All Files (*)")
        if not path:
            return
        self._current_path = path
        self.path_lbl.setText(path)
        self.reload_btn.setEnabled(True)
        self._load(path)

    def _reload(self):
        path = getattr(self, "_current_path", None)
        if path:
            self._load(path)

    def _load(self, path: str):
        try:
            col_idx = self.col_spin.value() - 1   # 0-based
            self._data = self._read_file(path, col_idx)
            self._populate_table()

            self.status_lbl.setText(
                f"{len(self._data)} module(s) loaded from column {self.col_spin.value()}.")
            self.status_lbl.setStyleSheet(f"color:{C_GREEN}; padding-top:6px; border:none;")
        except Exception as e:
            self.status_lbl.setText(f"Error: {e}")
            self.status_lbl.setStyleSheet(f"color:{C_ERRCLR}; padding-top:6px; border:none;")

    # ── file reading ──────────────────────────────────────────────────
    @staticmethod
    def _looks_like_header(value: str) -> bool:
        """Skip spreadsheet column-heading cells mistaken for module names."""
        lower = value.lower().strip()
        stem = os.path.splitext(lower)[0]
        headers = {
            "uut_name", "uut_file", "uut name", "uut file",
            "filename", "file name", "file_name", "c file", "c filename",
            "module", "modules", "name", "file", "source",
            "unit under test", ".c file",
        }
        return lower in headers or stem in headers

    def _read_file(self, path: str, col_idx: int):
        """
        Read a single column (col_idx, 0-based) from the spreadsheet.
        Both UUT_NAME (stem) and UUT_FILE (full .c name) are derived from
        whatever value is in that column.
        """
        raw_values = []

        if path.lower().endswith(".csv"):
            import csv
            with open(path, newline="", encoding="utf-8-sig") as f:
                for r in csv.reader(f):
                    if col_idx < len(r):
                        v = r[col_idx].strip()
                        if v:
                            raw_values.append(v)
        else:
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError(
                    "openpyxl not installed.\nRun: pip install openpyxl\nOr use a CSV file.")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb["Module Summary"] if "Module Summary" in wb.sheetnames else wb.active
            for row in ws.iter_rows(values_only=True):
                if col_idx < len(row) and row[col_idx] is not None:
                    v = str(row[col_idx]).strip()
                    if v:
                        raw_values.append(v)
            wb.close()

        # Derive (UUT_NAME, UUT_FILE) from each raw cell value
        rows = []
        for v in raw_values:
            if self._looks_like_header(v):
                continue
            # Ensure it ends with .c
            fname = v if v.lower().endswith(".c") else f"{v}.c"
            stem  = os.path.splitext(fname)[0]   # strip .c for UUT_NAME
            if self._looks_like_header(stem):
                continue
            rows.append((stem, fname))
        return rows

    # ── preview table ─────────────────────────────────────────────────
    def _populate_table(self):
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        for i, (name, file) in enumerate(self._data):
            self.table.insertRow(i)

            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk_item.setCheckState(Qt.Checked)
            chk_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 0, chk_item)

            for col, val in enumerate((str(i + 1), name, file), start=1):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.table.setItem(i, col, item)
        self.table.blockSignals(False)

        # Reset search filter and select-all to reflect the freshly loaded data
        self.search_edit.blockSignals(True)
        self.search_edit.clear()
        self.search_edit.blockSignals(False)
        for i in range(self.table.rowCount()):
            self.table.setRowHidden(i, False)

        self.select_all_chk.blockSignals(True)
        self.select_all_chk.setChecked(True)
        self.select_all_chk.blockSignals(False)

        self._update_selected_count()

    # ── checkbox / select-all handling ──────────────────────────────────
    def _checked_rows(self):
        """Row indices (matching self._data order) that are checked."""
        rows = []
        for i in range(self.table.rowCount()):
            item = self.table.item(i, 0)
            if item and item.checkState() == Qt.Checked:
                rows.append(i)
        return rows

    def _on_select_all_toggled(self, checked: bool):
        if self._syncing_select_all:
            return
        self._syncing_select_all = True
        self.table.blockSignals(True)
        for i in range(self.table.rowCount()):
            if self.table.isRowHidden(i):
                continue
            item = self.table.item(i, 0)
            if item:
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        self.table.blockSignals(False)
        self._syncing_select_all = False
        self._update_selected_count()

    def _on_item_changed(self, item):
        if item.column() != 0:
            return
        self._sync_select_all_state()
        self._update_selected_count()

    def _sync_select_all_state(self):
        total = checked = 0
        for i in range(self.table.rowCount()):
            if self.table.isRowHidden(i):
                continue
            item = self.table.item(i, 0)
            if item:
                total += 1
                if item.checkState() == Qt.Checked:
                    checked += 1
        self._syncing_select_all = True
        self.select_all_chk.setChecked(total > 0 and checked == total)
        self._syncing_select_all = False

    def _update_selected_count(self):
        total   = len(self._data)
        checked = len(self._checked_rows())
        if total:
            self.selected_count_lbl.setText(f"({checked} of {total} selected)")
        else:
            self.selected_count_lbl.setText("")

    # ── search / filter ──────────────────────────────────────────────
    def _filter_table(self, text: str):
        text = text.strip().lower()
        for i in range(self.table.rowCount()):
            if not text:
                self.table.setRowHidden(i, False)
                continue
            name_item = self.table.item(i, 2)
            file_item = self.table.item(i, 3)
            name_val  = name_item.text().lower() if name_item else ""
            file_val  = file_item.text().lower() if file_item else ""
            match = text in name_val or text in file_val
            self.table.setRowHidden(i, not match)
        self._sync_select_all_state()
        self._update_selected_count()

    # ── send to batch ─────────────────────────────────────────────────
    def _send_to_batch(self):
        if not self._data:
            QMessageBox.warning(self, "No Data", "Load an Excel/CSV file first.")
            return

        checked_rows = self._checked_rows()
        if not checked_rows:
            QMessageBox.warning(self, "Nothing Selected",
                "Check at least one .c module to send to UT Compilation.")
            return

        selected_data = [self._data[i] for i in checked_rows]

        self._batch_tab.imported_excel_path = getattr(self, "_current_path", "")
        self._batch_tab.set_modules(selected_data, replace=self.replace_chk.isChecked())
        self._switch_to_batch()
        QMessageBox.information(self, "Sent",
            f"{len(selected_data)} of {len(self._data)} module(s) sent to UT Compilation tab.")